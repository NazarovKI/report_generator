
# Программа Добавляет в xlsx картинки и подписи из имени файла
# Нужно:    # 1. изменить размер картинки # 2. дабавить подпись
import os
# import pillow
from openpyxl import Workbook
from openpyxl.drawing.image import Image
wb = Workbook()
ws = wb.active
os.chdir('C:\\Users\Пользователь\Desktop\Имущество\Храм\Иконы')
icons = os.listdir() # Создаёт список, содержащий имена файлов в рабочей папке
for i in icons: # Цикл, который
    img = Image(i) # Открывает Изображение из списка
    img.width = 70
    img.height = 81
    pos = 'D'+str(icons.index(i)+1) # Создаёт значение ячейки в которую  будет помещено изображение
    ws.add_image(img, pos) # Добавляет изображение img в ячейку pos 


icons = os.listdir() # Создаёт список, содержащий имена файлов в рабочей папке

for i in icons: # Цикл, который
    words = i.split('_')
    num = words[0]
    name = words[1].replace('jpg','')
    pos = 'B'+str(icons.index(i)+1) # Создаёт значение ячейки в которую  будет помещена надпись
    ws[pos] = name # Добавляет надпись рисунка, взятую из имени файла в ячейку pos
    pos = 'A'+str(icons.index(i)+1) # Создаёт значение ячейки в которую  будет помещена надпись
    ws[pos] = num # Добавляет номер рисунка, взятую из имени файла в ячейку pos

wb.save('icons.xlsx') # Сохраняет документ
